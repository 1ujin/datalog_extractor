#!/usr/bin/env python3
# -*- encoding:utf-8 -*-
"""
@Author : Lu Jin
@File   : app/util.py
@Time   : 2023/5/15
@Desc   : 工具类
"""

import sys
from decimal import Decimal
from openpyxl import styles
from openpyxl import utils
from openpyxl import Workbook
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidget

MIN_SIZE = -sys.maxsize - 1

LOOKUP_UNIT = {
    "a": "AMPERE", "ma": "AMPERE", "ua": "AMPERE", "na": "AMPERE", "pa": "AMPERE",
    "v": "VOLT", "mv": "VOLT", "uv": "VOLT", "nv": "VOLT", "pv": "VOLT",
    "w": "WATT", "mw": "WATT", "uw": "WATT", "nw": "WATT", "pw": "WATT",
    "hz": "HERTZ", "khz": "HERTZ", "mhz": "HERTZ", "ghz": "HERTZ",
}

UNIT = {
    "AMPERE":   {"A":  1e12, "mA":  1e9, "uA":  1e6, "nA":  1e3,  "pA": 1e0},
    "VOLT":     {"V":  1e12, "mV":  1e9, "uV":  1e6, "nV":  1e3,  "pV": 1e0},
    "WATT":     {"W":  1e12, "mW":  1e9, "uW":  1e6, "nW":  1e3,  "pW": 1e0},
    "HERTZ":    {"Hz": 1e0,  "kHz": 1e3, "MHz": 1e6, "GHz": 1e9},
}


def export_excel(table_widget, filename, progress=None, taskbar_progress=None):
    wb = Workbook()
    wb.remove(wb.active)
    for i in range(table_widget.count()):
        progress.setLabelText("正在导出（" + str(i + 1) + "/" + str(table_widget.count()) + ")")
        ws = wb.create_sheet()
        ws.title = table_widget.tabText(i)
        export_sheet(table_widget.widget(i), ws, 1, progress, taskbar_progress, table_widget.freeze_cell)
    wb.save(filename)


def export_sheet(table: QTableWidget, ws, header_row_cnt, progress=None, taskbar_progress=None, freeze_cell=None):
    max_row = table.rowCount()
    max_col = table.columnCount()
    total = max_col * max_row

    if progress:
        progress.setRange(0, total)
        progress.setValue(0)
    if sys.platform == "win32" and taskbar_progress is not None:
        taskbar_progress.setRange(0, total)
        taskbar_progress.setValue(0)

    # 测试值
    done = 0
    for col in range(0, max_col):
        col_letter = ws.cell(1, col + 1).column_letter
        max_length = ws.column_dimensions[col_letter].width
        it = iter(range(1, max_row))
        for row in it:
            if progress is not None and progress.wasCanceled():
                raise Exception("user canceled")
            item = table.item(row, col)
            if not item:
                done += 1
                if progress is not None:
                    progress.setValue(done)
                if sys.platform == "win32" and taskbar_progress is not None:
                    taskbar_progress.setValue(done)
                continue
            span = table.rowSpan(row, col)
            done += span
            cell = ws.cell(row + 1, col + 1)
            if len(item.text()) == 0:
                continue
            text = item.text().strip()
            if text[0] == "-":
                text = text[1:]
            if span > 1:
                cell.value = item.text().strip()
            elif text.find(".") > -1:
                integer, decimal = text.split(".", 1)
                if integer.isdigit() and decimal.isdigit():
                    cell.value = float(item.text().strip())
                else:
                    cell.value = item.text().strip()
            elif text.isdigit():
                cell.value = int(item.text().strip())
            else:
                cell.value = item.text().strip()
            rgb = item.background().color().rgb() % (256 * 256 * 256)
            if rgb != 0:
                cell.fill = styles.PatternFill(patternType="solid", fgColor=hex(rgb)[2:])
            max_length = max(max_length, len(str(cell.value)))
            if span > 1:
                ws.merge_cells(start_row=row + 1, start_column=col + 1, end_row=row + span, end_column=col + 1)
                for _ in range(span - 1):
                    next(it)
            cell.alignment = styles.alignment.Alignment(horizontal="center", vertical="center")
            text_alignment = item.textAlignment()
            if text_alignment & int(Qt.AlignLeft | Qt.AlignVCenter) == int(Qt.AlignLeft | Qt.AlignVCenter):
                cell.alignment = styles.alignment.Alignment(horizontal="left", vertical="center")
            elif text_alignment & int(Qt.AlignRight | Qt.AlignVCenter) == int(Qt.AlignRight | Qt.AlignVCenter):
                cell.alignment = styles.alignment.Alignment(horizontal="right", vertical="center")
            elif text_alignment & int(Qt.AlignHCenter | Qt.AlignTop) == int(Qt.AlignHCenter | Qt.AlignTop):
                cell.alignment = styles.alignment.Alignment(horizontal="center", vertical="top")
            if progress is not None:
                progress.setValue(done)
            if sys.platform == "win32" and taskbar_progress is not None:
                taskbar_progress.setValue(done)
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > max_length:
            # col_letter = [x[0].column_letter for x in ws.columns][col]
            ws.column_dimensions[col_letter].width = adjusted_width

    done = total - max_col
    if progress is not None:
        progress.setValue(done)
    if sys.platform == "win32" and taskbar_progress is not None:
        taskbar_progress.setValue(done)
    # 表头
    for row in range(header_row_cnt):
        it = iter(range(max_col))
        for col in it:
            item = table.item(row, col)
            if not item:
                done += 1
                if progress is not None:
                    progress.setValue(done)
                if sys.platform == "win32" and taskbar_progress is not None:
                    taskbar_progress.setValue(done)
                continue
            row_span = table.rowSpan(row, col)
            col_span = table.columnSpan(row, col)
            span = row_span * col_span
            done += span
            cell = ws.cell(row + 1, col + 1)
            cell.value = item.text()
            cell.alignment = styles.alignment.Alignment(horizontal="center", vertical="center")
            text_alignment = item.textAlignment()
            if text_alignment & int(Qt.AlignLeft | Qt.AlignVCenter) == int(Qt.AlignLeft | Qt.AlignVCenter):
                cell.alignment = styles.alignment.Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row + 1, start_column=col + 1, end_row=row + row_span, end_column=col + col_span)
            if col_span > 1:
                for _ in range(col_span - 1):
                    next(it)
            if progress is not None:
                progress.setValue(done)
            if sys.platform == "win32" and taskbar_progress is not None:
                taskbar_progress.setValue(done)

    # 左上角单元格
    row_span = table.rowSpan(0, 0)
    col_span = table.columnSpan(0, 0)
    ws.merge_cells(start_row=1, start_column=1, end_row=row_span, end_column=col_span)

    if progress is not None:
        progress.reset()
    if sys.platform == "win32" and taskbar_progress is not None:
        taskbar_progress.resume()
        taskbar_progress.reset()

    # 冻结
    if freeze_cell:
        ws.freeze_panes = utils.get_column_letter(freeze_cell[0]) + str(freeze_cell[1])

    # 隐藏
    for i in range(max_row):
        if table.isRowHidden(i):
            ws.row_dimensions[i + 1].hidden = True
    for i in range(max_col):
        if table.isColumnHidden(i):
            ws.column_dimensions[utils.get_column_letter(i + 1)].hidden = True


def isnumber(text):
    if not text and len(text) == 0:
        return False
    if text[0] == "-" or text[0] == "+":
        text = text[1:]
    values = text.split(".")
    if len(values) > 2:
        return False
    for val in values:
        if not val.isdigit():
            return False
    return True


def convert_decimal(val):
    if val == 0:
        return Decimal(0)
    elif val == val.to_integral():
        val = val.to_integral()
    else:
        val = val.normalize()
    return val


def convert_unit(val: Decimal, src_unit: str, tar_unit: str) -> Decimal:
    if not isinstance(val, Decimal):
        val = Decimal(val)
    res = val
    if tar_unit is None or len(tar_unit.strip()) == 0:
        return res
    for unit_dict in UNIT.values():
        for src_item in unit_dict.items():
            if src_unit.lower() == src_item[0].lower():
                res = res * Decimal(str(src_item[1]))
                for tar_item in unit_dict.items():
                    if tar_unit.lower() == tar_item[0].lower():
                        res = res / Decimal(str(tar_item[1]))
                        return res
    return val
